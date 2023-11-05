import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
from tempfile import NamedTemporaryFile
import os
# Function to clean header values
def clean_header_value(header_value):
    return ' '.join(header_value.split())

# Function to find headers in the worksheet and clean them
def find_headers(ws, expected_headers):
    for row in ws.iter_rows(values_only=True):
        cleaned_row = [clean_header_value(str(cell)) if cell is not None else '' for cell in row]
        headers = {cell: idx+1 for idx, cell in enumerate(cleaned_row) if clean_header_value(cell) in expected_headers}
        if len(headers) == len(expected_headers):
            return headers
    return None

# Define the mapping between the two sheets.
mapping = {
    'No': 'NO',
    'Description': 'DESCRIPTION',
    'Part Number': 'SKU NO',
    'Qty': 'QTY',
    'Partner Unit Price (RM)': 'UNIT PRICE (RM)',
    # 'Partner Total Price (RM)': 'TOTAL PRICE (RM)'
}

# Streamlit UI
st.title("Excel Data Transfer")
st.subheader("Select the source and destination Excel files")

# File uploader widgets
source_file = st.file_uploader("Choose a source file", type=['xlsx'])
destination_file = st.file_uploader("Choose a destination file", type=['xlsx'])

if source_file and destination_file:
    source_tab = st.text_input("Enter the source tab name:", "MBSB - 3 years")
    destination_tab = st.text_input("Enter the destination tab name:", "Sheet 1")

    if st.button('Transfer Data'):
        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_source, \
             NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_destination:

            # Write uploaded files to temp files
            tmp_source.write(source_file.getvalue())
            tmp_destination.write(destination_file.getvalue())

            # Load the workbooks
            wb1 = openpyxl.load_workbook(tmp_source.name)
            wb2 = openpyxl.load_workbook(tmp_destination.name)

            if source_tab in wb1.sheetnames and destination_tab in wb2.sheetnames:
                ws1 = wb1[source_tab]
                ws2 = wb2[destination_tab]

                try:
                    # Prepare the expected headers for source and destination
                    expected_source_headers = [clean_header_value(header) for header in mapping.keys()]
                    expected_dest_headers = [clean_header_value(header) for header in mapping.values()]

                    # Find the headers in the source and destination sheets
                    source_headers = find_headers(ws1, expected_source_headers)
                    dest_headers = find_headers(ws2, expected_dest_headers)

                    if source_headers and dest_headers:
                        # Initialize the row in ws2 where the data will start to be copied
                        dest_row_index = 21

                        # Iterate over the rows in ws1 starting from row 14 since row 13 is the header
                        for row in ws1.iter_rows(min_row=14, max_col=ws1.max_column, values_only=True):
                            # Iterate over the mapped fields and copy the data from ws1 to ws2
                            for source_header, dest_header in mapping.items():
                                # Get the corresponding column index for source and destination headers
                                source_col_index = source_headers[clean_header_value(source_header)]
                                dest_col_index = dest_headers[clean_header_value(dest_header)]

                                # Get the value from the source cell
                                source_value = row[source_col_index - 1]  # Adjust for zero index

                                # Write the value to the corresponding cell in the destination sheet
                                ws2.cell(row=dest_row_index, column=dest_col_index, value=source_value)

                            # Move to the next row in the destination sheet
                            dest_row_index += 1

                        # Save the modified workbook
                        wb2.save(tmp_destination.name)
                        wb2.close()
                        wb1.close()

                        # Download link
                        with open(tmp_destination.name, "rb") as fp:
                            btn = st.download_button(
                                label="Download Excel file",
                                data=fp,
                                file_name="modified_excel.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        st.success("Data transfer complete! Download the modified file.")
                    else:
                        st.error("Could not find all required headers in the source or destination sheet.")

                except Exception as e:
                    st.error(f"An error occurred: {e}")

            else:
                st.error("The specified tabs were not found in the uploaded files.")

            # Clean up temp files
            os.unlink(tmp_source.name)
            os.unlink(tmp_destination.name)

st.write("Upload both files and input tab names to enable the transfer button.")
